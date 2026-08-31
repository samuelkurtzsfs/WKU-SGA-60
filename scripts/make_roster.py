#!/usr/bin/env python3
"""Turn the roster into an Excel workbook.

    python3 scripts/build_roster.py     # first, writes site/roster*.csv
    python3 scripts/make_roster.py      # then, writes SGA-60-roster.xlsx

Reads site/roster.csv and site/roster-people.csv, which build_roster.py
compiles out of data/years.json, and adds the two things a spreadsheet needs
that a CSV does not: names split into first and last so the list can be sorted
and mail-merged, and formatting.

It deliberately derives nothing of its own. Everything here comes from the
roster the build already produces, so the workbook cannot drift from the site.
Run build_roster.py first, or this will use whatever it last wrote.
"""

import argparse
import csv
import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
ROSTER = ROOT / "site" / "roster.csv"
PEOPLE = ROOT / "site" / "roster-people.csv"

SUFFIXES = {"jr", "jr.", "sr", "sr.", "ii", "iii", "iv", "v"}
PARTICLES = {"van", "von", "de", "del", "della", "di", "da", "la", "le",
             "mac", "mc", "st", "st.", "o'"}

NICK_QUOTED = re.compile(r'"([^"]+)"|“([^”]+)”')
NICK_PAREN = re.compile(r"\(([^)]+)\)")


def split_name(name):
    """Split a recorded name into first / middle / last / suffix / known-as.

    Keeps the name people actually went by: Forrest (Bucky) Lanning and
    Edward "Eddie" Myers both keep Bucky and Eddie rather than losing them.
    """
    nickname = ""
    m = NICK_QUOTED.search(name)
    if m:
        nickname = (m.group(1) or m.group(2) or "").strip()
        name = NICK_QUOTED.sub(" ", name)
    m = NICK_PAREN.search(name)
    if m:
        nickname = nickname or m.group(1).strip()
        name = NICK_PAREN.sub(" ", name)
    name = " ".join(name.split())

    tokens = name.split()
    if not tokens:
        return "", "", "", "", nickname
    if len(tokens) == 1:
        return "", "", tokens[0], "", nickname

    suffix = ""
    if tokens[-1].strip(",").lower() in SUFFIXES and len(tokens) > 2:
        suffix = tokens[-1].strip(",")
        tokens = tokens[:-1]

    # A surname particle belongs with the surname, not the middle name.
    cut = len(tokens) - 1
    while cut > 1 and tokens[cut - 1].lower().strip(".") in PARTICLES:
        cut -= 1

    return tokens[0], " ".join(tokens[1:cut]), " ".join(tokens[cut:]), \
        suffix, nickname


# Plain English for the roster's own vocabulary. "Senate officer" and "Senate"
# are different things there: the first is the Speaker, the clerks and the
# class and constituency representatives, the second the rank and file.
BODY = {"Leader": "President or student regent",
        "Executive": "Executive officer",
        "Senate officer": "Senate officer",
        "Senate": "Senator",
        "Committee": "Committee chair",
        "Judicial": "Judicial branch"}

HEAD = PatternFill("solid", fgColor="0B0B0C")
HEADF = Font(color="FFFFFF", bold=True, size=10)


def sheet(wb, title, headers, data, widths, freeze="A2"):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font = HEAD, HEADF
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in data:
        ws.append(row)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = freeze
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    ws.row_dimensions[1].height = 30
    return ws


def read(path):
    if not path.is_file():
        sys.exit(f"{path} is missing. Run python3 scripts/build_roster.py first.")
    with path.open(newline="") as f:
        return list(csv.DictReader(f))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("-o", "--out", default=str(ROOT / "SGA-60-roster.xlsx"))
    args = ap.parse_args()

    terms = read(ROSTER)
    people = read(PEOPLE)

    wb = Workbook()
    wb.remove(wb.active)

    # ---- every term served
    rows = []
    for t in terms:
        first, middle, last, suffix, nick = split_name(t["person"])
        recorded = t.get("name_as_recorded", "")
        rows.append([
            last, first, middle, suffix, nick, t["person"],
            recorded if recorded != t["person"] else "",
            t["year"], (t["year"] or "")[:4],
            BODY.get(t.get("body", ""), t.get("body", "")),
            t.get("office", ""),
            "yes" if t.get("acting") else "",
            "yes" if t.get("also_regent") else "",
            t.get("name_verified", ""), t.get("plaque_term", ""),
            t.get("what_they_did", ""),
            t.get("source", ""), t.get("source_url", ""),
        ])
    rows.sort(key=lambda r: (str(r[8]), r[0].lower(), r[1].lower()))

    sheet(wb, "Every term served",
          ["Last name", "First name", "Middle", "Suffix", "Known as",
           "Full name", "Recorded in the source as", "Academic year",
           "Year began", "Branch", "Office held", "Acting",
           "Also held the regent seat", "Name verified", "Plaque reads",
           "What the archive says they did", "Source", "Source link"],
          rows,
          [18, 15, 12, 7, 12, 24, 24, 13, 11, 24, 34, 8, 15, 12, 12, 70, 34, 46])

    # ---- one row per person
    def tidy_offices(text):
        """The leader record and the executive record of a presidency are two
        rows in the archive and both reach the roster, so the office list can
        read "president (2023-24); President (2023-24)". Same office, same
        year, one mention."""
        seen, out = set(), []
        for part in (text or "").split(";"):
            part = part.strip()
            if not part:
                continue
            key = part.lower()
            if key in seen:
                continue
            seen.add(key)
            out.append(part[0].upper() + part[1:] if part else part)
        return "; ".join(out)

    prows = []
    for p in people:
        first, middle, last, suffix, nick = split_name(p["person"])
        prows.append([
            last, first, middle, suffix, nick, p["person"],
            p.get("also_recorded_as", ""),
            p.get("first_year", ""), p.get("last_year", ""),
            int(p["years_served"]) if str(p.get("years_served", "")).isdigit()
            else p.get("years_served", ""),
            p.get("years", ""),
            (lambda o: o[0].upper() + o[1:] if o else o)(
                p.get("most_senior_office", "")),
            tidy_offices(p.get("all_offices", "")),
            p.get("what_they_did", ""),
            p.get("sources", ""),
        ])
    prows.sort(key=lambda r: (r[0].lower(), r[1].lower()))

    sheet(wb, "People",
          ["Last name", "First name", "Middle", "Suffix", "Known as",
           "Full name", "Also recorded as", "First year", "Last year",
           "Years served", "Every year", "Most senior office",
           "Every office held", "What the archive says they did", "Sources"],
          prows,
          [18, 15, 12, 7, 12, 24, 26, 11, 11, 12, 30, 30, 60, 70, 40])

    # ---- what is in here, and how complete it is
    by_body = defaultdict(int)
    for t in terms:
        by_body[BODY.get(t.get("body", ""), t.get("body", ""))] += 1
    sen_years = defaultdict(int)
    for t in terms:
        if t.get("body") == "Senate":
            sen_years[t["year"]] += 1
    all_years = sorted({t["year"] for t in terms})
    no_senators = [y for y in all_years if y not in sen_years]
    thin = sorted(sen_years.items(), key=lambda kv: kv[1])[:5]
    with_narrative = sum(1 for t in terms if (t.get("what_they_did") or "").strip())

    notes = [
        ["What this is",
         "Everyone the SGA 60 archive records as holding an office in student "
         "government at Western Kentucky University, 1966-67 to 2026-27. The "
         "first sheet has one row per person per office per year, so somebody "
         "who served three years appears three times. The People sheet has "
         "one row each instead."],
        ["Where it comes from",
         "site/roster.csv and site/roster-people.csv, which the site build "
         "compiles out of data/years.json. To refresh this workbook after new "
         "research lands: python3 scripts/build_roster.py, then python3 "
         "scripts/make_roster.py."],
        ["", ""],
        ["Terms recorded", len(terms)],
        ["Distinct people", len(people)],
        ["Academic years covered", len(all_years)],
        ["Terms with an account of what the person did", with_narrative],
        ["", ""],
    ]
    for k in sorted(by_body, key=lambda k: -by_body[k]):
        notes.append([f"Terms recorded, {k.lower()}", by_body[k]])
    notes += [
        ["", ""],
        ["How complete this is",
         f"Senators are recorded for {len(sen_years)} of {len(all_years)} "
         f"years. The years with none at all are "
         f"{', '.join(no_senators) if no_senators else 'none'}. Even a year "
         "with senators is unlikely to have all of them: the roll was rebuilt "
         "from whichever minutes and rosters survive, and a meeting nobody "
         "minuted leaves no trace. Treat every year's count as a floor rather "
         "than a total. The thinnest are "
         + ", ".join(f"{y} ({n})" for y, n in thin) + "."],
        ["Names",
         "Full name is the spelling the archive settled on. Where a source "
         "spelled it differently, that spelling is kept beside it, so Christy "
         "Vogt and Christy Mollozzi stay visibly the same person. First and "
         "last name are split from the full name for sorting; a name someone "
         "actually went by, like Bucky or Eddie, is kept in its own column."],
        ["Empty cells",
         "An empty cell means the archive holds nothing for that field. It "
         "does not mean the thing is untrue. A thin record is a thin record, "
         "not a quiet year."],
    ]

    ws = sheet(wb, "Read me first", ["", ""], notes, [48, 104], freeze="A1")
    for r in range(1, ws.max_row + 1):
        ws.cell(row=r, column=1).font = Font(bold=True, size=10)
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True,
                                                       vertical="top")
    wb.move_sheet("Read me first", offset=-2)

    wb.save(args.out)
    print(f"wrote {args.out}")
    print(f"  {len(terms)} terms, {len(people)} people, "
          f"{len(all_years)} academic years")
    for k in sorted(by_body, key=lambda k: -by_body[k]):
        print(f"  {by_body[k]:5d}  {k}")


if __name__ == "__main__":
    main()
